import os
import time
import json
import base64
import asyncio
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from services.face_detector import detect_faces_4k_ultra
from services.aws_client import search_face_on_aws
from services.email_service import send_attendance_email
from services.attendance import parse_identity
from core.timezone_utils import get_time_str, get_date_str

class EventController:
    def __init__(self, broadcast_func=None):
        self.broadcast_func = broadcast_func

    async def broadcast_event(self, event_dict: dict):
        if self.broadcast_func:
            try:
                await self.broadcast_func(event_dict)
            except Exception as e:
                print(f"[EVENT WS ERROR] {e}")

    async def process_event_batch(self, event_id: str, event_name: str, event_date: str, event_dept: str, uploaded_files: list):
        """
        Processes batch of 4K/DSLR event photos:
        - High-density sliced SAHI face detection (zero pixel loss)
        - Sequential step-by-step progress telemetry over WebSocket
        - AWS Rekognition collection matching & cross-photo deduplication
        - Automated Excel compilation & email dispatch
        """
        print(f"\n[EVENT PIPELINE] 🚀 Starting 4K Event Ingestion: '{event_name}' (ID: {event_id}, {len(uploaded_files)} photos)")
        
        os.makedirs(f"static/event_photos/{event_id}", exist_ok=True)
        os.makedirs("reports", exist_ok=True)

        total_frames = len(uploaded_files)
        total_faces_detected = 0
        unique_attendees = {}  # key: roll_or_name -> dict
        frame_summaries = []

        await self.broadcast_event({
            "type": "event_started",
            "event_id": event_id,
            "event_name": event_name,
            "total_frames": total_frames,
            "message": f"Starting 4K ultra-resolution processing for '{event_name}' ({total_frames} photos)..."
        })

        for idx, file_item in enumerate(uploaded_files):
            frame_num = idx + 1
            filename = file_item["filename"]
            img_bytes = file_item["bytes"]

            # Save full 4K frame to disk
            local_path = f"static/event_photos/{event_id}/frame_{frame_num}_{filename}"
            try:
                with open(local_path, "wb") as f:
                    f.write(img_bytes)
            except Exception as e:
                print(f"Error saving event photo {filename}: {e}")

            # Telemetry Stage 1: Sliced 4K Detection Starting
            await self.broadcast_event({
                "type": "event_progress",
                "event_id": event_id,
                "frame_index": frame_num,
                "total_frames": total_frames,
                "filename": filename,
                "stage": "DETECTING",
                "message": f"Frame {frame_num}/{total_frames} [{filename}]: Slicing 4K canvas into high-res tiles (Zero Pixel Loss)..."
            })

            # Run Ultra-Res Sliced (SAHI) Detection in background thread
            start_detect = time.time()
            faces = await asyncio.to_thread(detect_faces_4k_ultra, img_bytes)
            detect_time_ms = int((time.time() - start_detect) * 1000)
            faces_count = len(faces)
            total_faces_detected += faces_count

            # Telemetry Stage 2: Detected Faces
            await self.broadcast_event({
                "type": "event_progress",
                "event_id": event_id,
                "frame_index": frame_num,
                "total_frames": total_frames,
                "filename": filename,
                "faces_found": faces_count,
                "detect_time_ms": detect_time_ms,
                "stage": "MATCHING",
                "message": f"Frame {frame_num}/{total_frames}: Found {faces_count} faces at native 4K optical density. Matching against AWS cloud vectors..."
            })

            # ── UPGRADE 3: Parallel AWS Batch Calls ──────────────────────────
            # All faces in a frame are matched simultaneously via asyncio.gather()
            # instead of one-at-a-time. For 21 faces this is ~21× faster AWS step.
            #
            # Each face also has TWO separate byte streams:
            #   face["bytes"]         → aligned+enhanced crop → sent to AWS
            #   face["display_bytes"] → natural padded crop   → shown in gallery
            # ─────────────────────────────────────────────────────────────────
            matched_in_this_frame = 0
            frame_crops = []

            # Build pre-broadcast crop records immediately (gallery shows instantly)
            pre_records = []
            valid_faces = []
            for face_idx, face in enumerate(faces):
                aws_bytes = face.get("bytes")        # aligned, for AWS
                disp_bytes = face.get("display_bytes") or aws_bytes  # natural crop, for gallery
                if not aws_bytes:
                    continue

                # Gallery image is the natural padded crop (display_bytes)
                disp_b64 = base64.b64encode(disp_bytes).decode("utf-8")

                record = {
                    "face_index": face_idx + 1,
                    "frame_num": frame_num,
                    "filename": filename,
                    "crop_b64": f"data:image/jpeg;base64,{disp_b64}",
                    "bbox": face.get("bbox", []),
                    "aligned": face.get("aligned", False),
                    "matched": False,
                    "name": "Scanning...",
                    "roll": "—",
                    "confidence": 0.0
                }
                pre_records.append(record)
                valid_faces.append((face_idx, aws_bytes))

                # Broadcast crop immediately so gallery fills in real-time
                await self.broadcast_event({
                    "type": "event_face_crop",
                    "event_id": event_id,
                    "frame_index": frame_num,
                    "face_index": face_idx + 1,
                    "total_faces_so_far": total_faces_detected,
                    "crop": record
                })

            # ── Parallel AWS matching for all faces in this frame ──
            async def _match_one(idx_in_list, aws_bytes_data):
                try:
                    res = await asyncio.to_thread(search_face_on_aws, aws_bytes_data)
                    if isinstance(res, list):
                        res = res[0] if res else {}
                    return idx_in_list, res
                except Exception as e:
                    print(f"[EVENT PARALLEL AWS] Face {idx_in_list} error: {e}")
                    return idx_in_list, {}

            aws_tasks = [_match_one(i, aws_b) for i, aws_b in enumerate(
                [vf[1] for vf in valid_faces]
            )]
            aws_results = await asyncio.gather(*aws_tasks)

            # Process AWS results and update records
            for list_idx, search_res in aws_results:
                record = pre_records[list_idx]
                frame_crops.append(record)

                if isinstance(search_res, dict) and search_res.get("match"):
                    matched_in_this_frame += 1
                    raw_id = search_res.get("identity", "Unknown")
                    name, roll = parse_identity(raw_id)
                    conf = search_res.get("confidence", 95.0)

                    record["matched"] = True
                    record["name"] = name
                    record["roll"] = roll
                    record["confidence"] = round(conf, 1)

                    key = f"{roll}_{name}".strip("_")
                    if key not in unique_attendees:
                        unique_attendees[key] = {
                            "name": name,
                            "roll_number": roll,
                            "confidence": round(conf, 1),
                            "first_seen_frame": frame_num,
                            "seen_count": 1,
                            "verified_time": get_time_str(),
                            "photo": record["crop_b64"]
                        }
                    else:
                        unique_attendees[key]["seen_count"] += 1
                        unique_attendees[key]["confidence"] = max(
                            unique_attendees[key]["confidence"], round(conf, 1)
                        )

                    # Re-broadcast with match result so gallery card updates to green
                    await self.broadcast_event({
                        "type": "event_face_matched",
                        "event_id": event_id,
                        "frame_index": frame_num,
                        "face_index": record["face_index"],
                        "crop": record
                    })

            frame_summaries.append({
                "frame": frame_num,
                "filename": filename,
                "faces_detected": faces_count,
                "matched_count": matched_in_this_frame,
                "detect_time_ms": detect_time_ms,
                "crops": frame_crops
            })

            # Telemetry Stage 3: Frame Complete
            await self.broadcast_event({
                "type": "event_frame_completed",
                "event_id": event_id,
                "frame_index": frame_num,
                "total_frames": total_frames,
                "filename": filename,
                "faces_detected": faces_count,
                "matched_count": matched_in_this_frame,
                "current_unique_total": len(unique_attendees),
                "total_faces_detected_cumulative": total_faces_detected,
                "message": f"Frame {frame_num}/{total_frames} complete ({faces_count} faces, {matched_in_this_frame} matched). Unique Attendees: {len(unique_attendees)}"
            })

        # --- 4. COMPILE EXCEL WORKBOOK REPORT ---
        report_filename = f"event_{event_id}.xlsx"
        report_path = os.path.join("reports", report_filename)
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Event Attendance Ledger"

            # Header Styling
            header_fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
            header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            title_font = Font(name="Arial", size=14, bold=True, color="0F172A")
            data_font = Font(name="Arial", size=10)
            center_align = Alignment(horizontal="center", vertical="center")
            left_align = Alignment(horizontal="left", vertical="center")
            border = Border(
                left=Side(style='thin', color='E2E8F0'),
                right=Side(style='thin', color='E2E8F0'),
                top=Side(style='thin', color='E2E8F0'),
                bottom=Side(style='thin', color='E2E8F0')
            )

            # Metadata Title Rows
            ws.merge_cells("A1:F1")
            ws["A1"] = f"🎓 Anyiiiiie.AI — {event_name.upper()}"
            ws["A1"].font = title_font
            ws["A1"].alignment = left_align

            ws["A2"] = "Event Date:"
            ws["B2"] = event_date or get_date_str()
            ws["C2"] = "Department / Hall:"
            ws["D2"] = event_dept or "Main Auditorium"
            ws["E2"] = "Photos Ingested:"
            ws["F2"] = f"{total_frames} (4K Ultra-Res)"

            ws["A3"] = "Total Faces Scanned:"
            ws["B3"] = total_faces_detected
            ws["C3"] = "Verified Attendees:"
            ws["D3"] = len(unique_attendees)
            ws["E3"] = "Generated:"
            ws["F3"] = f"{get_time_str()} IST"

            for r in range(2, 4):
                for col in ["A", "B", "C", "D", "E", "F"]:
                    ws[f"{col}{r}"].font = Font(name="Arial", size=10, bold=(col in ["A", "C", "E"]))

            # Table Header
            headers = ["S.No", "Roll Number / ID", "Faculty / Student Name", "Recognition Confidence", "Photos Present In", "Verification Status"]
            ws.append([]) # Row 4 empty
            ws.append(headers) # Row 5

            for col_num in range(1, 7):
                cell = ws.cell(row=5, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align

            # Data Rows
            sorted_attendees = sorted(unique_attendees.values(), key=lambda x: x["name"])
            for idx, att in enumerate(sorted_attendees):
                row_idx = 6 + idx
                ws.append([
                    idx + 1,
                    att["roll_number"],
                    att["name"],
                    f"{att['confidence']}%",
                    f"{att['seen_count']} photo(s)",
                    "VERIFIED PRESENT ✓"
                ])
                for c in range(1, 7):
                    cell = ws.cell(row=row_idx, column=c)
                    cell.font = data_font
                    cell.border = border
                    cell.alignment = center_align if c in [1, 2, 4, 5, 6] else left_align

            # Auto Column Width
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

            wb.save(report_path)
            print(f"[EVENT PIPELINE] ✅ Excel report compiled: {report_path}")
        except Exception as e:
            print(f"[EVENT PIPELINE] Error compiling Excel report: {e}")

        # --- 5. AUTOMATED EMAIL SUMMARY VIA RESEND ---
        email_sent = False
        try:
            attendee_list_for_email = list(unique_attendees.values())
            session_meta = {
                "id": event_id,
                "name": event_name,
                "duration_minutes": "Event Session",
                "attendees": attendee_list_for_email,
                "raw_frames": [f"static/event_photos/{event_id}/frame_{s['frame']}_{s['filename']}" for s in frame_summaries[:6]]
            }
            email_sent = send_attendance_email(session_meta, report_path)
        except Exception as e:
            print(f"[EVENT PIPELINE] Email dispatch error: {e}")

        # --- 6. FINAL BROADCAST TO DASHBOARD ---
        final_payload = {
            "type": "event_completed",
            "event_id": event_id,
            "event_name": event_name,
            "event_date": event_date,
            "total_frames": total_frames,
            "total_faces_detected": total_faces_detected,
            "unique_attendees_count": len(unique_attendees),
            "attendees": list(unique_attendees.values()),
            "frame_summaries": frame_summaries,
            "report_url": f"/api/download_report/{report_filename}",
            "email_sent": email_sent,
            "message": f"🎉 4K Event Ingestion Complete! {len(unique_attendees)} Unique Attendees Verified across {total_frames} photos."
        }

        await self.broadcast_event(final_payload)
        print(f"[EVENT PIPELINE] 🎉 4K Event Pipeline Finished for '{event_name}': {len(unique_attendees)} verified attendees.")
        return final_payload
